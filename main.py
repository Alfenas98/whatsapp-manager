import time
import json
import random
import threading
import asyncio
import sys
import os
from pathlib import Path
from datetime import datetime

import openpyxl
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.prompt import Prompt, Confirm
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.live import Live
from rich.layout import Layout
from rich.text import Text
from rich import box

console = Console()

PROGRESS_FILE = "progresso.json"
LOG_FILE = "log_envios.txt"

# ─────────────────────────────────────────────
# UTILITÁRIOS
# ─────────────────────────────────────────────

def salvar_progresso(dados: dict):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def carregar_progresso() -> dict:
    if Path(PROGRESS_FILE).exists():
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def log(mensagem: str):
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    linha = f"[{timestamp}] {mensagem}"
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(linha + "\n")

def limpar_tela():
    os.system("cls" if os.name == "nt" else "clear")

def cabecalho():
    limpar_tela()
    console.print(Panel(
        "[bold green]WHATSAPP MANAGER v1.0[/bold green]\n"
        "[dim]Extrator de Contatos + Disparador de Mensagens[/dim]",
        border_style="green",
        expand=False
    ))
    console.print()

# ─────────────────────────────────────────────
# MÓDULO 1 — EXTRATOR DE CONTATOS
# ─────────────────────────────────────────────

def extrair_contatos():
    cabecalho()
    console.print(Panel("[bold cyan]MÓDULO 1 — EXTRATOR DE CONTATOS[/bold cyan]", border_style="cyan"))
    console.print()

    console.print("[yellow]⚠️  Este módulo abre o WhatsApp Web no Chrome.[/yellow]")
    console.print("[dim]Você precisará escanear o QR Code na primeira vez.[/dim]")
    console.print()

    nome_grupo = Prompt.ask("[bold]Nome do arquivo de saída (sem extensão)[/bold]", default="contatos_grupo")
    pasta_downloads = Path.home() / "Downloads"
    pasta_downloads.mkdir(exist_ok=True)
    arquivo_saida = str(pasta_downloads / f"{nome_grupo}.xlsx")

    if not Confirm.ask(f"\nSalvar contatos em [green]{arquivo_saida}[/green]?"):
        return

    console.print("\n[dim]Iniciando navegador...[/dim]")

    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            # Caminhos padrão do Chrome no Windows
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
            ]
            chrome_exe = next((p for p in chrome_paths if os.path.exists(p)), None)

            if not chrome_exe:
                console.print("[red]❌ Chrome não encontrado! Instale o Google Chrome e tente novamente.[/red]")
                return

            browser = p.chromium.launch(
                headless=False,
                executable_path=chrome_exe,
                args=["--start-maximized"]
            )
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 800}
            )
            page = context.new_page()
            page.goto("https://web.whatsapp.com")

            console.print("\n[bold yellow]📱 Escaneie o QR Code no navegador que abriu.[/bold yellow]")
            console.print("[dim]Aguardando login...[/dim]")

            # Aguarda o login (aparecimento da barra lateral)
            page.wait_for_selector('div[aria-label="Lista de conversas"]', timeout=120000)
            console.print("[green]✅ Login realizado![/green]")
            console.print()
            console.print("[bold]Agora abra o grupo desejado no WhatsApp Web e pressione ENTER aqui.[/bold]")
            input()

            console.print("[dim]Aguardando carregamento do grupo...[/dim]")
            time.sleep(2)

            # Abre painel de participantes
            console.print("[dim]Abrindo lista de participantes...[/dim]")

            try:
                # Clica no cabeçalho do grupo para abrir info
                page.click('header[data-testid="conversation-header"]')
                time.sleep(1)

                # Rola a lista de participantes para carregar todos
                console.print("[dim]Carregando participantes (pode demorar)...[/dim]")

                numeros = set()
                tentativas = 0
                max_tentativas = 60  # 60 rolagens

                while tentativas < max_tentativas:
                    # Busca elementos de participantes
                    elementos = page.query_selector_all('div[data-testid="cell-frame-container"]')

                    for el in elementos:
                        try:
                            texto = el.inner_text()
                            linhas = texto.strip().split("\n")
                            for linha in linhas:
                                linha = linha.strip().replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
                                # Detecta padrão de número brasileiro
                                if linha.startswith("+55") and len(linha) >= 13:
                                    numeros.add(linha)
                                elif linha.startswith("55") and len(linha) >= 12:
                                    numeros.add("+" + linha)
                        except:
                            pass

                    # Rola para baixo dentro do painel
                    painel = page.query_selector('div[data-testid="drawer-right"]')
                    if painel:
                        painel.evaluate("el => el.scrollTop += 300")
                    else:
                        page.keyboard.press("End")

                    time.sleep(0.5)
                    tentativas += 1

                    if tentativas % 10 == 0:
                        console.print(f"[dim]Números encontrados até agora: {len(numeros)}...[/dim]")

            except Exception as e:
                console.print(f"[red]Erro ao navegar no grupo: {e}[/red]")
                browser.close()
                return

            browser.close()

            if not numeros:
                console.print("[red]❌ Nenhum número encontrado. Verifique se o grupo estava aberto.[/red]")
                return

            # Salva no Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contatos"
            ws.append(["numero", "enviado", "erro"])

            for numero in sorted(numeros):
                ws.append([numero, "NAO", ""])

            # Formata colunas
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 30

            wb.save(arquivo_saida)

            console.print()
            console.print(Panel(
                f"[bold green]✅ Extração concluída![/bold green]\n"
                f"Números encontrados: [bold]{len(numeros)}[/bold]\n"
                f"Arquivo salvo: [bold]{arquivo_saida}[/bold]",
                border_style="green"
            ))

    except Exception as e:
        console.print(f"[red]Erro: {e}[/red]")

    console.print()
    input("Pressione ENTER para voltar ao menu...")

# ─────────────────────────────────────────────
# MÓDULO 2 — DISPARADOR DE MENSAGENS
# ─────────────────────────────────────────────

def configurar_mensagens() -> list:
    cabecalho()
    console.print(Panel("[bold cyan]CONFIGURAR MENSAGENS[/bold cyan]", border_style="cyan"))
    console.print("[dim]Digite as 3 mensagens. Serão sorteadas aleatoriamente no envio.[/dim]\n")

    mensagens = []
    for i in range(1, 4):
        console.print(f"[bold yellow]Mensagem {i}:[/bold yellow]")
        console.print("[dim](Digite a mensagem e pressione ENTER duas vezes para confirmar)[/dim]")
        linhas = []
        while True:
            linha = input()
            if linha == "" and linhas:
                break
            linhas.append(linha)
        mensagem = "\n".join(linhas)
        mensagens.append(mensagem)
        console.print(f"[green]✅ Mensagem {i} salva![/green]\n")

    return mensagens

def configurar_intervalo() -> dict:
    cabecalho()
    console.print(Panel("[bold cyan]CONFIGURAR INTERVALOS[/bold cyan]", border_style="cyan"))
    console.print()

    console.print("[bold]Intervalo entre mensagens (segundos):[/bold]")
    minimo = int(Prompt.ask("  Mínimo", default="15"))
    maximo = int(Prompt.ask("  Máximo", default="45"))

    console.print()
    console.print("[bold]Tamanho do lote:[/bold]")
    lote = int(Prompt.ask("  Mensagens por lote", default="50"))

    console.print()
    console.print("[bold]Pausa entre lotes (minutos):[/bold]")
    pausa = int(Prompt.ask("  Minutos de pausa", default="10"))

    return {
        "intervalo_min": minimo,
        "intervalo_max": maximo,
        "lote": lote,
        "pausa_minutos": pausa
    }

def disparar_mensagens():
    cabecalho()
    console.print(Panel("[bold cyan]MÓDULO 2 — DISPARADOR DE MENSAGENS[/bold cyan]", border_style="cyan"))
    console.print()

    # Verifica progresso anterior
    progresso = carregar_progresso()
    retomar = False

    if progresso:
        console.print(f"[yellow]⚠️  Progresso anterior encontrado:[/yellow]")
        console.print(f"   Arquivo: [bold]{progresso.get('arquivo', '?')}[/bold]")
        console.print(f"   Enviados: [bold]{progresso.get('enviados', 0)}[/bold]")
        console.print(f"   Pendentes: [bold]{progresso.get('pendentes', 0)}[/bold]")
        console.print()
        retomar = Confirm.ask("Deseja retomar de onde parou?")
        console.print()

    # Seleciona arquivo
    if retomar:
        arquivo = progresso["arquivo"]
    else:
        arquivo = Prompt.ask("[bold]Caminho da planilha Excel[/bold]", default="contatos_grupo.xlsx")

    if not Path(arquivo).exists():
        console.print(f"[red]❌ Arquivo '{arquivo}' não encontrado![/red]")
        input("\nPressione ENTER para voltar...")
        return

    # Carrega contatos
    wb = openpyxl.load_workbook(arquivo)
    ws = wb.active

    contatos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        numero = str(row[0]).strip() if row[0] else ""
        enviado = str(row[1]).strip().upper() if row[1] else "NAO"
        if numero and (not retomar or enviado != "SIM"):
            contatos.append(numero)

    if not contatos:
        console.print("[yellow]⚠️  Nenhum contato pendente encontrado.[/yellow]")
        input("\nPressione ENTER para voltar...")
        return

    console.print(f"[green]✅ {len(contatos)} contatos carregados.[/green]\n")

    # Configura mensagens
    mensagens = configurar_mensagens()

    # Configura intervalos
    config = configurar_intervalo()

    cabecalho()
    console.print(Panel("[bold cyan]RESUMO DO DISPARO[/bold cyan]", border_style="cyan"))

    tabela = Table(box=box.SIMPLE)
    tabela.add_column("Configuração", style="dim")
    tabela.add_column("Valor", style="bold green")
    tabela.add_row("Total de contatos", str(len(contatos)))
    tabela.add_row("Intervalo entre msgs", f"{config['intervalo_min']}–{config['intervalo_max']} seg")
    tabela.add_row("Tamanho do lote", str(config['lote']))
    tabela.add_row("Pausa entre lotes", f"{config['pausa_minutos']} min")
    tabela.add_row("Mensagens configuradas", "3 (sorteio aleatório)")
    console.print(tabela)
    console.print()

    if not Confirm.ask("[bold yellow]Iniciar envio agora?[/bold yellow]"):
        return

    console.print("\n[dim]Iniciando navegador...[/dim]")

    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
                os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe"),
            ]
            chrome_exe = next((cp for cp in chrome_paths if os.path.exists(cp)), None)

            if not chrome_exe:
                console.print("[red]Chrome nao encontrado! Instale o Google Chrome.[/red]")
                return

            browser = p.chromium.launch(
                headless=False,
                executable_path=chrome_exe,
                args=["--start-maximized"]
            )
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 800}
            )
            page = context.new_page()
            page.goto("https://web.whatsapp.com")

            console.print("\n[bold yellow]📱 Escaneie o QR Code no navegador que abriu.[/bold yellow]")
            console.print("[dim]Aguardando login...[/dim]")
            page.wait_for_selector('div[aria-label="Lista de conversas"]', timeout=120000)
            console.print("[green]✅ Login realizado! Iniciando disparos...[/green]\n")
            time.sleep(2)

            enviados = 0
            falhas = 0
            pausado = False

            # Atualiza coluna "enviado" no Excel
            def marcar_enviado(numero: str, status: str, erro: str = ""):
                for row in ws.iter_rows(min_row=2):
                    if str(row[0].value).strip() == numero:
                        row[1].value = status
                        row[2].value = erro
                        break
                wb.save(arquivo)

            for idx, numero in enumerate(contatos):
                # Verifica se é hora de pausar entre lotes
                if idx > 0 and idx % config["lote"] == 0:
                    pausa_seg = config["pausa_minutos"] * 60
                    console.print(f"\n[yellow]☕ Lote concluído. Pausando por {config['pausa_minutos']} minutos...[/yellow]")
                    log(f"PAUSA entre lotes. Enviados até agora: {enviados}")

                    for s in range(pausa_seg, 0, -1):
                        console.print(f"[dim]  Retomando em {s}s...[/dim]", end="\r")
                        time.sleep(1)
                    console.print()

                # Sorteia mensagem
                mensagem = random.choice(mensagens)

                # Sorteia intervalo
                intervalo = random.randint(config["intervalo_min"], config["intervalo_max"])

                try:
                    # Monta URL do WhatsApp
                    from urllib.parse import quote
                    url = f"https://web.whatsapp.com/send?phone={numero}&text={quote(mensagem)}"
                    page.goto(url)

                    # Aguarda campo de texto carregar
                    page.wait_for_selector('div[data-testid="conversation-compose-box-input"]', timeout=20000)
                    time.sleep(1.5)

                    # Envia
                    page.keyboard.press("Enter")
                    time.sleep(1)

                    enviados += 1
                    marcar_enviado(numero, "SIM")
                    log(f"ENVIADO → {numero}")

                    # Salva progresso
                    salvar_progresso({
                        "arquivo": arquivo,
                        "enviados": enviados,
                        "pendentes": len(contatos) - enviados - falhas,
                        "falhas": falhas,
                        "ultimo_numero": numero
                    })

                    console.print(
                        f"[green]✅[/green] [{idx+1}/{len(contatos)}] {numero} "
                        f"[dim]| msg {mensagens.index(mensagem)+1} | próximo em {intervalo}s[/dim]"
                    )

                except Exception as e:
                    falhas += 1
                    erro_msg = str(e)[:80]
                    marcar_enviado(numero, "ERRO", erro_msg)
                    log(f"FALHA → {numero} | {erro_msg}")
                    console.print(f"[red]❌[/red] [{idx+1}/{len(contatos)}] {numero} [dim]→ {erro_msg}[/dim]")

                # Aguarda intervalo
                time.sleep(intervalo)

            browser.close()

            console.print()
            console.print(Panel(
                f"[bold green]🎉 Disparo concluído![/bold green]\n"
                f"✅ Enviados: [bold]{enviados}[/bold]\n"
                f"❌ Falhas:   [bold]{falhas}[/bold]\n"
                f"📋 Log salvo em: [bold]{LOG_FILE}[/bold]",
                border_style="green"
            ))

            # Limpa progresso ao finalizar com sucesso
            if Path(PROGRESS_FILE).exists():
                os.remove(PROGRESS_FILE)

    except Exception as e:
        console.print(f"[red]Erro crítico: {e}[/red]")

    console.print()
    input("Pressione ENTER para voltar ao menu...")

# ─────────────────────────────────────────────
# MENU PRINCIPAL
# ─────────────────────────────────────────────

def menu_principal():
    while True:
        cabecalho()

        # Verifica progresso pendente
        progresso = carregar_progresso()
        if progresso:
            console.print(Panel(
                f"[yellow]⚠️  Há um disparo pausado!\n"
                f"Enviados: {progresso.get('enviados', 0)} | "
                f"Pendentes: {progresso.get('pendentes', 0)}[/yellow]",
                border_style="yellow"
            ))
            console.print()

        tabela = Table(show_header=False, box=box.SIMPLE, padding=(0, 2))
        tabela.add_column("Opção", style="bold green", width=4)
        tabela.add_column("Descrição")
        tabela.add_row("1", "Extrair contatos de grupo")
        tabela.add_row("2", "Disparar mensagens")
        tabela.add_row("3", "Ver log de envios")
        tabela.add_row("0", "Sair")
        console.print(tabela)
        console.print()

        opcao = Prompt.ask("[bold]Escolha uma opção[/bold]", choices=["0", "1", "2", "3"])

        if opcao == "1":
            extrair_contatos()
        elif opcao == "2":
            disparar_mensagens()
        elif opcao == "3":
            cabecalho()
            if Path(LOG_FILE).exists():
                console.print(Panel(f"[bold]LOG DE ENVIOS — {LOG_FILE}[/bold]", border_style="dim"))
                with open(LOG_FILE, "r", encoding="utf-8") as f:
                    linhas = f.readlines()
                for linha in linhas[-50:]:  # últimas 50 linhas
                    if "ENVIADO" in linha:
                        console.print(f"[green]{linha.strip()}[/green]")
                    elif "FALHA" in linha:
                        console.print(f"[red]{linha.strip()}[/red]")
                    else:
                        console.print(f"[dim]{linha.strip()}[/dim]")
            else:
                console.print("[dim]Nenhum log encontrado ainda.[/dim]")
            console.print()
            input("Pressione ENTER para voltar...")
        elif opcao == "0":
            console.print("\n[dim]Encerrando...[/dim]\n")
            sys.exit(0)

if __name__ == "__main__":
    menu_principal()
